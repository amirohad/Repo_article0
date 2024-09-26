# -*- coding: utf-8 -*-
"""
Created on Sun Dec 12 11:03:31 2021
useful functions for data extraction in twine experiment
# 1 get length in pix from image
# 2 get x,y pix position from image
# 3 draw line and get it's length
# 4 square distance
# 5 draw lines with live update and return line length
# 6 array -> dictionary -> data frame -> excel
# 7 get exp details
# 8 append to list
# 9 sum multiple lines in photo to one length
# 10 adjust length
# 11 3d distance between 2 points
# 12 simple animated plot
# 13 change names of imgs in bending file to names in txt file by order
# 14 resize img keep aspect ratio
# 15 filter img by hsv
# 16 get avg centerline for thresholded shape
# 16b get centerline with linear interpolation of nans and using median
# 17 show two images one above the other
# 18 zoom/crop - preserves pixel size in cm
# 19 threshold filter for image
# 20 difference between lines (remove start bias?)
# 21 open normalized window
# 22 get cn periods from csv list data
# 23 from img folder, copy to dst file imgs at constant interval
# 24 get cn times
# 25 ROI center
# 26 distance between 2 ROI centers
# 27 align 2 photos
# 28 find closest value in list to number
# 29 permutation differnce test
# 30 non-parametric tests
# 31 remove outliers for time series
# 32 find cicrle radius and center from 3 points
# 33 tolerant mean
# 34 angle between 2 lines
# 35 mark lines on image
# 36 CN from track
# 37 fit to sine
# 38 remove nans and []
# 39 deny_confirm
# 40 place normal window on external monitor at certain size
# 41 plot definitions
# 42 analytic functions
# 43 linear fit with errors
# 44 get image resolution (can change other data type)
# 45 write pandas dataframe to excel
@author: Amir

"""
#%% imports
import cv2
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import math as m
import scipy
import scipy.stats
import re
import matplotlib.pyplot as plt
import matplotlib.lines as mlines
import matplotlib.animation as animation

from openpyxl import load_workbook
import shutil # for copy and other high level operations
import os,glob # for listing files in folder
import re
import timeit
import PIL # image functions
import pickle # for saving data
# import imutils # image utilities
# import itertools # Functions creating iterators for efficient looping
import random
import inspect
from tqdm import tqdm
# from progress_bar import InitBar # progress bar
print('hi')
#%% 1 get length of section in image
def get_pix_length(filename,length_name):
    img = cv2.imread(filename) # load file
    cv2.namedWindow(length_name,cv2.WINDOW_NORMAL)
    roi = cv2.selectROI(length_name, img,showCrosshair=True, 
                fromCenter=False) # select roi- returns [x0,y0,width,height]
    # cv2.waitKey(0)
    cv2.destroyAllWindows()
    dist = int((np.sqrt(roi[-2]**2+roi[-1]**2)))
    return dist

#%% 2 get y coor. of point in image
def get_ypix(filename,object_name):
    img = cv2.imread(filename) # load file
    cv2.namedWindow("Mark object y position of"+object_name,cv2.WINDOW_NORMAL)
    roi = cv2.selectROI("Mark object y position of"+object_name, img,showCrosshair=True, fromCenter=False) # select roi- returns [x0,y0,width,height]
    # cv2.waitKey(0)
    cv2.destroyAllWindows()
    ypos = int(roi[1]) # get y position of uppper left corner
    return ypos

#%% 3 draw line and get it's length
def draw_line(event,x,y,flags,params): # mouse callback function
    ''' get input from mouse as event, with coordinates x,y. draw line,
    from initial click to current mouse position
    params=[window_name,img,xy,drawing]'''
    global x0,y0
    window_name = params[0]
    img = params[1]
    xy = params[2]
    drawing = params[3]
    # linewidth = int(img.shape[0]/1000)
    linewidth = 1
    if event == cv2.EVENT_LBUTTONDOWN:
        drawing = True
        x0,y0 = x,y
    elif event == cv2.EVENT_MOUSEMOVE:
        if drawing == True:
            img_c = img.copy()
            cv2.line(img_c,(x0,y0),(x,y),(0,255,0),linewidth)
            cv2.imshow(window_name,img_c)
    elif event == cv2.EVENT_LBUTTONUP:
        drawing = False
        xy.append([x0,y0,x,y]) # update line coordinates each buttonup
        cv2.line(img,(x0,y0),(x,y),(0,255,0),linewidth)
#%% 4 square distance
def distance_2d(p1p2): #p1p2 = [x0,y0,x,y]
    (x1, y1,x2,y2) = p1p2
    return round(np.sqrt((x1-x2)**2 + (y1-y2)**2))
#%% 5 draw lines with live update and return line length
def live_line(filename,window_name,xy,xpic=500,ypix=0):
    if type(filename)==str: o_img = cv2.imread(filename) # load file
    else: o_img = filename
    window_name += ', d to del line, q to exit'
    cv2.namedWindow(window_name,cv2.WINDOW_NORMAL)
    cv2.moveWindow(window_name, xpic,ypix)
    # c_img = o_img.copy()
    # x0,y0 = -1,-1
    drawing = False # true if mouse is pressed
    stay = True
    while stay:
        cv2.setMouseCallback(window_name,draw_line,[window_name,o_img,xy,drawing])
        cv2.imshow(window_name,o_img)
        if cv2.waitKey(1) & 0xFF == ord("q"): stay = False # exit with q
        elif cv2.waitKey(1) & 0xFF == ord("d"): del xy[-1]# delete last line with d
    cv2.destroyAllWindows()
    return len(xy)
#%% 6 array -> dictionary -> data frame -> excel
def arr_to_excel(filepath,d_array,columns,writer='xlsxwriter'):
    writer = pd.ExcelWriter(filepath, engine = 'xlsxwriter')

    rows_list = [] #insert data here as dictionaries
    for row in d_array:
        dict1 = {}
        i=0
        for column in columns:
            # convert input row into dictionary format
            # key = col_name
            if np.size(row)>1:dict1.update({column:row[i]})
            else:dict1.update({column:row})
            i+=1

            # append dictionary row to list
        rows_list.append(dict1)
    # convert list to dataframe
    df = pd.DataFrame(rows_list)
    # save to excel
    df.to_excel(writer)
    writer.save()
#%% 7 get exp details
# filepath = r'C:\Users\Amir\Documents\PHD\Experiments\Force Measurements\Exp2_Pendulum\init_contact\081_top.JPG'
def exp_details(filepath,exp_num,event_num,view):
    '''filepath, exp_num, event_num,view'''
    # name = '098_1_top.JPG'
    exp_num.append(re.findall('[0-9][0-9][0-9]',filepath)[0])
    if re.findall('_[0-9]_',filepath):
        event_num.append(re.findall('_[0-9]_',filepath)[0][1])
    else: event_num.append(1)
    vw = re.findall('_[a-z]+.JPG',filepath)[0][-8:-3]
    regex = re.compile('[,_\.!?]') # what to remove
    view.append(regex.sub('', vw))

#%% 8 append to list
def append_to_list(lists,strt,copies): #index says how far back to copy
    '''lists, start position, # of copies'''
    for l in lists:
        for i in range(copies):
            if strt=='NA' or l[-1]==[]:
                l.append('NA')
            else:
                for row in l[strt:]:
                    # l.append(l[indx])
                    l.append(row)
#%% 9 sum multiple lines in photo to one length
def multiline_dist(file,title=[],all_dist=False):
    '''press q to finish, d to delete last line, returns sum of lines'''
    line = []
    live_line(file,title,line)
    sumlen = 0
    for l in line: # sum all lines drawn
        sumlen += distance_2d(l) # sum over all lines
    if all_dist == False: return sumlen
    else: return [distance_2d(l) for l in line]
#%% 10 adjust length
def adjust_len(a,b,choose=[]): # add strt,end?
    '''a,b are diff length, make same length.
        choose = a, to make b same length as a'''
    la=len(a)
    lb=len(b)
    if la>lb:
        if choose is None or choose is a:
            b = np.pad(b,[0,la-lb])
        elif np.array_equal(choose, b):
            a = a[:lb]
    elif la<lb:
        if choose is None or choose is b:
            a = np.pad(a,[0,lb-la])
        elif np.array_equal(choose, a):
            b = b[:la]

    return a,b
#%% 11 nd distance between 2 points
def distance_nd(vec1,vec2):
    '''(x,y,z,)1 , (x,y,z)2 or other dimension vector
    x,y,w,h-> diagonals are: (x,y), (x+w,y+h)'''
    return np.sqrt(sum([(v1i-v2i)**2 for v1i,v2i in zip(vec1,vec2)]))

#%% 12 simple animated plot

def simp_ani(xx,yy,ftitle,xtitle=[],ytitle=[],xlims=[],ylims=[],delay=100):
    '''xdata,ydata,xtitle,ytitle,xlims,ylims. Always save returned
        output to a variable'''
    fig = plt.figure()
    ax = plt.subplot(1,1,1)

    def init():
        ax.clear()
        plt.xlabel(xtitle)
        plt.ylabel(ytitle)
        plt.title(ftitle)
        if xlims: ax.set_xlim(xlims)
        if ylims: ax.set_ylim(ylims)

    def update(i):
        ax.plot(xx[:i+1],yy[:i+1],'b')



    return ax,animation.FuncAnimation(fig, update,frames=len(xx),init_func=init,interval=delay)

# # example
# x1=np.arange(0,10.5,0.5)
# y=np.multiply(x1,x1)
# ani= simp_ani(x1,y,'example','x','x^2',[0,100],[0,100],100)
#%% 13 change names of imgs in bending file to names in txt file by order
def swap_names(path):
    '''change names of imgs in bending file to names in txt file(in same folder)'''
    file_lst = os.listdir(path) # path = folder path
    for i in range(len(file_lst)):
        names = open(path+r'\img_names.txt','r').readlines()
        names = [name[:-1] for name in names]
        if file_lst[i][-1]=='G':
            new_name = names[i][3:]+'.JPG' # start from 4th character in the name row
            full_old = path +'\\'+ file_lst[i] # old
            full_new_name = path + "\\" + new_name # new
            if full_new_name not in file_lst: # check didnt change already
                os.rename(full_old, full_new_name) # rename in same folder
                # print(full_old, full_new_name)
    print('done')
# uncomment for local use:
# base = r'C:\Users\Amir\Documents\PHD\Experiments\Force Measurements\Exp2_Pendulum\Measurements'
# exp = 119
# path = base+'\\' + str(exp)+r'\Bending' # input
# name_lst = swap_names (path)
#%% 14 resize img keep aspect ratio
def ResizeWithAspectRatio(image, width=None, height=None, inter=cv2.INTER_AREA):
    '''resize img with aspect ration, can set width/height/non. changes pixel info!'''
    dim = None
    (h, w) = image.shape[:2]

    if width is None and height is None:
        return image
    if width is None:
        r = height / float(h)
        dim = (int(w * r), height)
    else:
        r = width / float(w)
        dim = (width, int(h * r))

    return cv2.resize(image, dim, interpolation=inter)
#%% 15 filter img by hsv
def nothing(x):
    pass
def hsv_filter(img):
    '''color filter with track bars, return mask created by choosen hsv'''
    # Create a window
    win_name = 'hsv filter, q to quit'
    normal_window(win_name,img)
    # cv2.namedWindow(win_name,cv2.WINDOW_NORMAL)
    # cv2.imshow(win_name,img)
    cv2.moveWindow(win_name, 300,30)  # Move it to (500,30)

    # Create trackbars for color change
    # Hue is from 0-179 for Opencv
    cv2.createTrackbar('HMin', win_name, 0, 179, nothing)
    cv2.createTrackbar('SMin', win_name, 0, 255, nothing)
    cv2.createTrackbar('VMin', win_name, 0, 255, nothing)
    cv2.createTrackbar('HMax', win_name, 0, 179, nothing)
    cv2.createTrackbar('SMax', win_name, 0, 255, nothing)
    cv2.createTrackbar('VMax', win_name, 0, 255, nothing)

    # Set default value for Max HSV trackbars
    cv2.setTrackbarPos('HMax', win_name, 179)
    cv2.setTrackbarPos('SMax', win_name, 255)
    cv2.setTrackbarPos('VMax', win_name, 255)

    # Initialize HSV min/max values
    hMin = sMin = vMin = hMax = sMax = vMax = 0
    phMin = psMin = pvMin = phMax = psMax = pvMax = 0
    stay = True
    while stay:
        # Get current positions of all trackbars
        hMin = cv2.getTrackbarPos('HMin', win_name)
        sMin = cv2.getTrackbarPos('SMin', win_name)
        vMin = cv2.getTrackbarPos('VMin', win_name)
        hMax = cv2.getTrackbarPos('HMax', win_name)
        sMax = cv2.getTrackbarPos('SMax', win_name)
        vMax = cv2.getTrackbarPos('VMax', win_name)

        # Set minimum and maximum HSV values to display
        lower = np.array([hMin, sMin, vMin])
        upper = np.array([hMax, sMax, vMax])

        # Convert to HSV format and color threshold
        hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
        mask = cv2.inRange(hsv, lower, upper)
        result = cv2.bitwise_and(img, img, mask=mask)

        # Print if there is a change in HSV value
        if((phMin != hMin) | (psMin != sMin) | (pvMin != vMin) | (phMax != hMax) | (psMax != sMax) | (pvMax != vMax) ):
            print("(hMin = %d , sMin = %d, vMin = %d), (hMax = %d , sMax = %d, vMax = %d)" % (hMin , sMin , vMin, hMax, sMax , vMax))
            phMin = hMin
            psMin = sMin
            pvMin = vMin
            phMax = hMax
            psMax = sMax
            pvMax = vMax

        # Display result image
        cv2.imshow(win_name, result)
        if cv2.waitKey(10) & 0xFF == ord('q'):
            stay = False
    cv2.destroyAllWindows()

    return lower,upper
#%% 16 get centerline for thresholded shape
def get_centerline(thresh):
    '''on a thresholded img, returns the avg index of 'on'(white) pixel
        for each column(x)'''
    start = timeit.default_timer()
    place = []
    for i in range(np.shape(thresh)[1]): # for each row(!)[column?!]
        mask = np.where(thresh[:,i]==255)[0]
        if not mask.size>0: # if no pixel on column, take last position as value
            if i>0: place[-1]=place[-2]
            else: place.append(np.mean(np.shape(thresh)[0])) # mean
        if i>0:
            if np.isnan(place[-1]):
                if place[-2]: place[-1]=place[-2]
                else: place[-1]=np.mean(np.shape(thresh)[0]) # mean
        place.append(np.mean(mask))

    stop = timeit.default_timer()
    print('Time: ', stop - start)
    return place

#%% 16b get centerline with linear interpolation of nans and using median
def get_centerlineb(thresh):
    '''on a thresholded img, find nans, return the median along each column
    then linearly interpolate nans'''
    start = timeit.default_timer()
    center = []
    for i in range(np.shape(thresh)[1]):
        mask = np.where(thresh[:,i]==255)[0]
        if not mask.size>0: # if no pixel on column, write nan
                center.append(np.nan)
                continue # next column

        # get centerline with median
        center.append(np.median(mask))
    # interpolate over nan positions
    center = np.array(center)
    x = np.linspace(1,len(center),len(center))
    center_interp = np.interp(x,x[~np.isnan(center)],center[~np.isnan(center)])

    stop = timeit.default_timer()
    print('Time: ', stop - start)
    return center_interp
#%% 17 show two images one above the other
def show_both(im1,im2):
    '''show two images one above the other'''
    vis = np.concatenate((im1, im2), axis=0)
    normal_window('both', vis)

#%% 18 zoom/crop - preserves pixel size in cm
def zoom(filename,roi=[], zoom_target=[]):
    '''get filename for img to zoom in on, zoom target name optional.
        returns zoomed in image'''
    img = cv2.imread(filename)
    if zoom_target: pass
    else: zoom_target = 'zoom'
    if not roi:
        cv2.namedWindow(zoom_target,cv2.WINDOW_NORMAL)
        cv2.moveWindow(zoom_target, 500, 200)
        roi = cv2.selectROI(zoom_target,img)
        cv2.destroyWindow(zoom_target)
    else: pass
    # ROI: (x,y,w,h)
    return img[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]
#%% 19 threshold filter for image
def filter_thresh(img,hsv_lims,kernel_size,opening=False,show=False,reverse=False):
    ''' get img, filter with given mask, output thresholded img(binary) '''
    threshold = 137
    maxval = 255
    lower,upper = hsv_lims

    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    mask = cv2.inRange(hsv, lower, upper)
    img_hsv = cv2.bitwise_and(img, img, mask=mask)
    imgray = cv2.cvtColor(img_hsv,cv2.COLOR_BGR2GRAY) # convert to grayscale

    # threshold - > closing(regular or reverse)

    # Otsu's thresholding after Gaussian filtering
    gauss_kernel = (41,41)
    blur = cv2.GaussianBlur(imgray,gauss_kernel,0)
    Otsu_ret,Otsu_thresh = cv2.threshold(blur,0,255,cv2.THRESH_BINARY+cv2.THRESH_OTSU)

    # closing morphology
    a = kernel_size # filter box edge size
    close_kernel = np.ones((int(a),int(a)),np.uint8) # filter shape
    closing = cv2.morphologyEx(Otsu_thresh, cv2.MORPH_CLOSE, close_kernel)
    thresh = closing
    # Global thresholding
    #ret, threshc = cv2.threshold(closing,int(threshold/2),maxval,0) # find threshold of image-> convert to binary

    if reverse:
        close_kernel = np.ones((int(a),int(a)),np.uint8) # filter shape
        closing = cv2.morphologyEx(imgray, cv2.MORPH_CLOSE, close_kernel)
        # ret, threshr = cv2.threshold(closing,threshold,maxval,cv2.THRESH_BINARY_INV) # find inverse threshold of image-> convert to binary
        thresh = closing

    if opening:
        open_kernel = np.ones((int(a),int(a)),np.uint8)
        mask_open=cv2.morphologyEx(thresh,cv2.MORPH_OPEN, open_kernel)
        # ret, thresho = cv2.threshold(mask_open,int(threshold),maxval,0)
        thresh = mask_open

    # present result
    if show:
        pos = [1900,100]
        if opening: title = 'threshold opening'
        elif reverse: title = 'threshold reverse'
        else: title = 'threshold Otsu'
        normal_window(title,thresh)
        cv2.moveWindow(title, pos[0], pos[1])

    return thresh
#%% 20 difference between lines (remove start bias)
def line_diff(line1,line2):
    '''find difference for each x between 2 lines, without bias (subtract first element difference from all)'''
    # find difference between free and loaded centerlines (minus bias)
    # return [line2[k]-line1[k]-(line2[0]-line1[0])
    #                 for k in range(min(len(line1),len(line2)))]

    # try: abs diff
    # find absolute difference between free and loaded centerlines (minus bias)
    return [abs(line2[k]-line1[k]-(line2[0]-line1[0]))
                    for k in range(min(len(line1),len(line2)))]
#%% 21 open normalized window
def normal_window(win_name,img):
    cv2.namedWindow(win_name,cv2.WINDOW_NORMAL)
    cv2.imshow(win_name,img)
    cv2.moveWindow(win_name, 0,0)
#%% 22 get cn periods from csv list
def get_cn_times(path):
    frames = []
    for file in glob.glob(os.path.join(path, '*.JPG')): # for each event:
        if re.findall('\d{4,5}',file):
            frames.append(int(re.findall('\d{4,5}',file)[0]))
    cn_periods = [(t - s) for s, t in zip(frames, frames[1:])]
    return cn_periods,np.average(cn_periods)/2

# test =r'C:\Users\Amir\Documents\PHD\Experiments\Force Measurements\Exp2_Pendulum\Measurements\65\Side\cn'
# frames,avg_Tcn=get_cn_times(test)
# print(avg_Tcn)
#%% 23 from an img folder, copy imgs to dst file at constant file interval
def copy_from_list (src,dst,interval,first=[]):
    '''get src (source) folder, send copies to dst (destination) folder
    at given interval starting at first (given or 0th) file'''
    file_lst = os.listdir(src) # path = folder path
    i=0
    for file in file_lst:
        if re.findall('DSC_',file_lst[0]):
            if not first: first = int(file_lst[0][-8:][:-4])
            number = int(re.findall('DSC_\d{4,5}',file)[0][4:])
            full_old = src +"\\"+ file
            full_new = dst + "\\DSC_" + str(number)+".JPG"# new
            print (number)
        else:
            first=0
            number = i
            full_old = src +"\\"+ file
            full_new = dst +"\\"+ file# new
            i+=1
        if (number-first)%interval==0:
            print(full_old)
            print(full_new)
            shutil.copy2(full_old,full_new)
            print(number)
# uncomment for local use:
# first = 0
# interval = 20
# # input
# src = r"\\132.66.44.238\AmirOhad\PhD\Experiments\Angled Support\Exp1_Inclined\Measurements\60\14-exp136-plant1\Helda_upsideview\plant1\cropped"
# # output
# dst = src + r"14_up_croppedskip20"
# if not os.path.exists(dst):
#        os.makedirs(dst)
# copy_from_list(src,dst,interval,first)
# print ('done')
#%% 24 get cn times from cn imgs folder
def get_Tcn(CNfolder,df,i):
    '''cnfolder is where the cn images are, df is data frame of events
    i is the row within data frame'''
    try:
        file_lst = os.listdir(CNfolder)
        frame = []
        for file in file_lst:
            # if re.findall('DSC_\d{4,5}',file):
            #     frame.append(int(re.findall('DSC_\d{4,5}',file)[0][4:]))
            if re.findall('\d{4,5}',file):
                frame.append(int(re.findall('\d{4,5}',file)[0]))
        frame = sorted(frame)
        diff = abs(np.subtract(frame[1:],frame[:-1]))
        T_cn = np.divide(diff,2)
        avgT_cn = np.mean(T_cn)
        # print(T_cn,avgT_cn)
        return T_cn,avgT_cn
    except:
        singleT = float(df.at[i,'C.N_time(minutes)'])
        return singleT,singleT
# internal use:
# CNfolder = r'C:\Users\Amir\Documents\PHD\Experiments\Force Measurements\Exp2_Pendulum\Measurements\22\Side\cn'
# aa = get_Tcn(CNfolder,data_panda,2)
# print(aa)
#%% 25 ROI center
def ROIcenter(img,window_name=None):
    '''gets an image and returns center of selected ROI'''

    if not window_name:
        window_name = 'frame interest point in center of roi'
    print('mark '+window_name)
    cv2.namedWindow(window_name,cv2.WINDOW_NORMAL)
    cv2.imshow(window_name,img)
    x,y,w,h = cv2.selectROI(window_name,img, showCrosshair=True)
    cv2.destroyAllWindows()
    return (int(x + w / 2), int(y + h / 2))
#%% 26 distance between 2 ROI centers
def ROI_dist(roi1,roi2,axis=0):
    '''get 2 ROIs and return the distance along request axis.
    axis='x'':x distance, axis='y': y distance '''
    # ROI: (x,y,w,h)
    xy1 = ROIcenter(roi1)
    xy2 = ROIcenter(roi2)
    if axis==0:
        return distance_2d(xy1[0],xy1[1],xy2[0],xy2[1])
    elif axis=='x':
        return abs(xy1[1]-xy2[0])
    elif axis=='y':
        return abs(xy1[1]-xy2[1])
#%% 27 align 2 photos
def align_images(image,template,maxFeatures=500,keepPercent=0.2,debug=False):
    '''image- to align, same as template with slight shift, maxF- max on
    points to consider for alignment, keepPercent- what percent of point matches
    to keep(reduce noise),debug- show/dont matching points'''
    # convert imgs to grayscale
    imgGray = cv2.cvtColor(image,cv2.COLOR_BGR2GRAY)
    tmpGray = cv2.cvtColor(template,cv2.COLOR_BGR2GRAY)
    # use ORB to get keypoints and get binary local invariant features
    orb = cv2.ORB_create(maxFeatures)
    (kpsA,descsA) = orb.detectAndCompute(imgGray,None)
    (kpsB,descsB) = orb.detectAndCompute(tmpGray,None)
    # match the features
    method = cv2.DESCRIPTOR_MATCHER_BRUTEFORCE_HAMMING
    matcher = cv2.DescriptorMatcher_create(method)
    matches = matcher.match(descsA,descsB,None)

    # sort the matches by distnace(smaller=more similar)
    matches = sorted(matches,key=lambda x:x.distance)

    #keep top matches
    keep = int(len(matches)*(keepPercent))
    matches = matches[:keep]

    # if debug- show matched points
    if debug:
        matchedVis = cv2.drawMatches(image,kpsA,template,kpsB,matches,None)
        matchedVis = imutils.resize(matchedVis,width=1000)
        cv2.imshow("matched Keypoints",matchedVis)
        cv2.waitKey(0)

    # variables for keypoint coordinates from top matches
    ptsA = np.zeros((len(matches),2),dtype="float")
    ptsB = np.zeros((len(matches),2),dtype="float")

    # loop top matches and map keypoints
    for (i,k) in enumerate(matches):
        ptsA[i] = kpsA[k.queryIdx].pt
        ptsB[i] = kpsB[k.trainIdx].pt

    # compute homography matric for matched points
    (H,mask) = cv2.findHomography(ptsA,ptsB,method=cv2.RANSAC)
    # use homography matric to align images
    (h,w) = template.shape[:2]
    aligned = cv2.warpPerspective(image,H,(w,h))
    # return aligned img
    return aligned
#%% 28 find index of closest value in a list to input number
def closest(lst, K):

     lst = np.asarray(lst) # convert to array
     indx = (np.abs(lst - K)).argmin() # find minimal absolute distance
     return indx
#%% 29 permutation differnce test
def permutation_test(var1,var2,n_reps,x_range,var_name):
    all_vars =list(itertools.compress(var1, ~np.isnan(var1)))\
                +list(itertools.compress(var2, ~np.isnan(var2)))
    shuffled_differences = []  # An empty list to store the differences
    def difference_in_means(combined_list,n):
        """ Split shuffled combind group into two, return mean difference"""
        group_A = combined_list[:n]
        group_B = combined_list[n:]
        return abs(np.mean(group_B) - np.mean(group_A)) # try absolute value

    grpA = list(itertools.compress(var1, ~np.isnan(var1)))
    observed_difference = difference_in_means(all_vars,len(grpA))

    n_greater_equal = 0
    # pbar = InitBar()
    for i in tqdm(range(n_reps)):
        # if i%(n_reps/100)==0: pbar(100*i/n_reps)
        random.shuffle(all_vars)
        new_difference = difference_in_means(all_vars,len(grpA))
        # Collect the new mean by adding to the end of the list
        shuffled_differences.append(new_difference)
        if shuffled_differences[i] >= observed_difference:
            n_greater_equal = n_greater_equal + 1

    pd.DataFrame(shuffled_differences).plot.hist(
        stacked=False, bins=100, figsize=(10, 6),
        range=x_range,density=False,alpha=0.7,label= 'random permutations')
    pval  = n_greater_equal/n_reps
    plt.axvline(observed_difference,color='r',
                label = f"{n_reps} permutations, pvalue = {pval:E}")
    # label=[f"{n_reps} permutations, pvalue = {n_greater_equal/n_reps:.6f}"]
    plt.legend()
    plt.xlabel(var_name, fontsize=16)
    plt.ylabel('frequency', fontsize=16)
    return pval
#%% 30 non-parametric tests
def non_parametric(df1,df2,print = False):
    '''compare two pd.dataframe distributions,
        returns [ttest_pval,ks_pval,mises_pval.pvalue,manwhiteney_pval]'''
    #t-test
    df1 = df1.dropna()[df1.columns[0]]
    df2 = df2.dropna()[df2.columns[0]]
    stat, ttest_pval = scipy.stats.ttest_ind(df1,df2)


    # kolmogorov-smirnov test
    stat, ks_pval =  scipy.stats.kstest(df1, df2)


    # cramer von mises 2 sample test
    mises_pval =  scipy.stats.cramervonmises_2samp(df1, df2)


    # mann-whitney u-test
    stat, manwhiteney_pval = scipy.stats.mannwhitneyu(df1, df2)
    if print:
        print(f"cramer von-mises test: statistic={mises_pval.statistic:.2E},\
              p-value={mises_pval.pvalue:6f}")
        print(f"ks-test: statistic={stat:.4f}, p-value={ks_pval:.2E}")
        print(f"t-test: statistic={stat:.4f}, p-value={ttest_pval:.2E}")
        print(f" Mann–Whitney U Test: statistic={stat:.2E}, p-value={manwhiteney_pval:.2E}")
    return ttest_pval,ks_pval,mises_pval.pvalue,manwhiteney_pval
#%% 31 remove outliers for time series
def find_outliers(f,x,cut_off):
    '''f- function values, x- spacing vector'''
    grad = np.gradient(f,x)
    outliers = []
    # dont start from first element?
    for i in range(len(f)): # find outliers via gradient exceeding cutoff.
        if abs(grad[i])>cut_off:
            outliers.append(i)
    return outliers

def remove_outliers_loop(f,cutoff=1,plots=0,loop=3):
    '''check for data with gradient exceeding 'cutoff', repeat for 'loop' times'''
    outliers = [] # save outliers
    f1 = f.copy() # copy to not cover original array
    x = np.arange(len(f1)) # evenly spaced vector for interpolation
    for i in range(loop):
        outliers.append(find_outliers(f1,x,cut_off=cutoff))
        if len(outliers)>0:   # if there are outliers
            # interpolate curve without the outliers
            f1 = np.interp(x,np.delete(x,outliers),np.delete(f1,outliers))
            outliers = []
        else: return f
    return f1


def remove_outliers(f,cutoff=1,plots=0):
    f1 = f.copy() # copy to not cover original array
    # x = np.arange(0,len(f1)*30,30) # use actual t?
    x = np.arange(len(f1)) # or arbitrary t
    grad = np.gradient(f1,x) # calc gradient
    outliers = []
    for i in range(1,len(f1)-1): # find outliers via gradient exceeding cutoff
        if abs(grad[i])>cutoff:
            outliers.append(i)
    if len(outliers)>0:   # if there are outliers
    # interpolate curve without the outliers
        f_clean = np.interp(x,np.delete(x,outliers),np.delete(f1,outliers))
    else: return f

    if plots==1:
        print(f"outliers={outliers}")
        fig,ax = plt.subplots(1,2)
        ax[0].plot(x,f,'b*',markersize=3,label='data')
        ax[0].plot(np.array(x)[outliers],np.array(f)[outliers],'xr',label='outliers')
        ax[0].plot(x,f_clean,'k',label='cleaned')
        ax[0].legend()
        ax[1].plot(np.gradient(f,x),'b')
        ax[1].plot(np.gradient(f1,x),'--r')
        plt.figure()
        plt.plot(np.delete(x,outliers),np.delete(f1,outliers),'x',label='removed outliers')

    return f_clean
#%% 32 find cicrle radius and center from 3 points
def circle_from_3_points(z1:complex, z2:complex, z3:complex) -> tuple[complex, float]:
    '''insert (x1+y1j,x2+x2j,x3+y3j)'''
    if (z1 == z2) or (z2 == z3) or (z3 == z1):
        raise ValueError(f'Duplicate points: {z1}, {z2}, {z3}')

    w = (z3 - z1)/(z2 - z1)

    # You should use a small tolerance rather than 0 for floating point comparisons
    if abs(w.imag) <= 0:
        raise ValueError(f'Points are collinear: {z1}, {z2}, {z3}')

    c = (z2 - z1)*(w - abs(w)**2)/(2j*w.imag) + z1;  # Simplified denominator
    r = abs(z1 - c);

    return c, r
#%% 33 tolerant mean
def tolerant_mean(arrs):
    lens = [len(i) for i in arrs]
    arr = np.ma.empty((np.max(lens),len(arrs)))
    arr.mask = True
    for idx, l in enumerate(arrs):
        arr[:len(l),idx] = l
    return arr.mean(axis = -1), arr.std(axis=-1)

# y, error = tolerant_mean(list_of_ys_diff_len)
# ax.plot(np.arange(len(y))+1, y, color='green')
#%% 34 angle between 2 lines
def angle_2lines(line1, line2):
    # Calculate the angle between two lines in degrees
    angle = m.atan2(line2[1][1] - line2[0][1],
                    line2[1][0] - line2[0][0]) - m.atan2(line1[1][1] - line1[0][1],
                                                         line1[1][0] - line1[0][0])
    return m.degrees(angle)
#%% 35 mark line on image
def mark_line(image_path):
    image = cv2.imread(image_path)

    # Load the image
    if image is None:
        print("Error: Could not load image.")
        return

    zoom_im = zoom(image_path)
    cv2.imshow('zoomed in- mark-line',zoom_im)
    line1 = cv2.selectROI("Image", zoom_im, showCrosshair=False)
    cv2.destroyWindow("Image")

    return line1
# mark_line(r"\\132.66.44.238\AmirOhad\PhD\Experiments\Force Measurements\Exp2_Pendulum\Measurements\81\Top\1\DSC_1207.JPG")
#%% 36 CN from track
def CN_periods(track_file,center):
    '''get tip coordiantes from track file, get center from input(?).
    identify when tip crosses line connecting initial tip position and center.
    can check line crossed above and below center point, compare rates?'''

    borderline = (center,track_file[0])

#%% 37 fit sine
def fit_sin(tt, yy,guess_freq):
    '''Fit sin to the input time sequence, and return fitting parameters "amp",
        "omega", "phase"(?), "offset", "freq", "period" and "fitfunc"'''
    tt = np.array(tt)
    yy = np.array(yy)
    guess_amp = max (yy)
    guess_offset = 0
    # guess_phase = 0
    guess = np.array([guess_amp, 2*np.pi*guess_freq, guess_offset])

    popt, pcov = scipy.optimize.curve_fit(sinfunc, tt, yy,
                  p0=guess, bounds=([0, 0, 0],[np.inf,np.inf,0.1]))
    A, w, c = popt
    f = w/(2.*np.pi)

    # calc residuals
    residuals = yy- sinfunc(tt, *popt)
    # residual sum of squares (SSE)
    ss_res = np.sum(residuals**2)
    # get total sum of squares (ss_tot):
    ss_tot = np.sum((yy-np.mean(yy))**2)
    # calc R^2
    r_squared = 1 - (ss_res / ss_tot)
    # calc root mean square error RMSE:
    k = 3
    rmse = np.sqrt(ss_res/(len(yy)-k))
    # S is the square root of the MSE. And MSE = SS Error/DF error (N-k-1)
    S = np.sqrt(ss_res/(len(yy)-k-1))

    fitfunc = lambda t: sinfunc(t, A, w, c)
    return {"amp": A, "omega": w, "offset": c, "freq": f,
            "period": 1./f, "fitfunc": fitfunc, "maxcov": np.max(pcov),
            "rawres": (guess,popt,pcov),"r_squared":r_squared,"rmse":rmse,"S":S}
#%% 38 remove nans and []
def remove_nans(x):
    '''input a list to remove its nans and [], returns list without them'''
    return [a for a  in x if not
          (np.isnan(a) or (isinstance(a, list) and not np.array(a).size>0) or a == 0)]
#%% 39 deny_confirm
def deny_confirm():
    '''press 0 to deny or 1 to confirm'''
    # Wait for key press
    print('to confirm press 1, to deny press 0')
    while True:
        key = cv2.waitKey(0) & 0xFF
        # deny
        if key == ord('0'):
            return 0
        # confirm
        elif key == ord('1'):
            return 1
#%% 40 place normal window on external monitor at certain size
def move_window_to_monitor(resolution=[1920, 1080],xy_percent=[0.8,0.6],
                      window_name='Normal window'):
    # Get the screen resolution
    screen_width, screen_height = resolution[0], resolution[1] # use actual resolution of screen

    # Set relative size of window (e.g. 80% of screen width and 60% of screen height)
    window_width = int(xy_percent[0] * screen_width)
    window_height = int(xy_percent[1] * screen_height)

    # Calculate the position to center the window on the screen
    x_position = resolution[0]
    y_position = int((screen_height - window_height) / 2)

    # Resize the window to the specified size
    cv2.resizeWindow(window_name, window_width, window_height)

    # Move the window to the calculated position on the screen
    cv2.moveWindow(window_name, x_position, y_position)
    return
#%% 41. plot definitions
def plot_definitions(fig=[],ax=[],xlabel='x',ylabel='y',fs=18,
                     title=False,suptitle=False,xrange=[],yrange=[],
                     xtick=[],ytick=[],sharex=False,sharey=False,legend=False):
    '''plot definitions. if var not inputed, default used. returns: none'''
    # new fig
    if not ax:
        fig,ax = plt.subplots(1,1)
        print('new axis created')
    else: pass
    # shared axis labels
    if sharex: fig.supxlabel(xlabel,fontsize=fs+2)
    else: ax.set_xlabel(xlabel,fontsize=fs)
    if sharey: fig.supylabel(ylabel,fontsize=fs)
    else: ax.set_ylabel(ylabel,fontsize=fs)
    #titles
    if suptitle: fig.suptitle(suptitle,fontsize=fs+4)
    if title: ax.set_title(title,fontsize=fs+2)
    # ranges
    if xrange: ax.set_xlim(xrange)
    if yrange: ax.set_ylim(yrange)
    if legend: ax.legend()
    return

# plot_definitions()

#%% 42 analytic functions
def sinfunc(t, A, w, c):  return A * (np.sin(w*t)) + c
def wAcoswtfunc(t,B,w,c): return B * (np.cos(w*t)) + c
def logfunc(x, A, c):  return A * (np.log(x)) + c
def linfunc(x,a,b): return a*x + b

# Core fitting function (purely numerical)
def expfunc_core(x, a, b):
    return a * np.exp(b * x)
# Wrapper function for generating the string representation
def expfunc(x, a, b, to_print=False):
    if to_print:
        return rf'$f(x) = {a:.2f}e^{{{b:.2f}x}}$'
    else: return expfunc_core(x, a, b)

def expfunc_test(x,a,b): 
    return a*np.exp(b*x)

def polyfunc(x,an): 
    '''n is the degree of the polynomial'''
    # an = np.array(an)?
    for i in range(len(an)):
        y = an[i]*x**i
    return sum(y)
def gaussfunc(x, a, b, c): return a * np.exp(-b * (x - c)**2)

def powerfunc(x, a, b, c): return a * x**b + c

#%% 43 fit with errors
def fit_w_err(ax,x,dx,y,dy,fit_func = linfunc, data_color = 'blue',
              fit_color = 'red',add_legend = False, legend_loc='best',data_alpha=0.5):
    '''Fit data to the provided function and plot with error bars.

    Parameters:
    ax : matplotlib axis object
    x, dx : array-like, data and errors in x
    y, dy : array-like, data and errors in y
    fit_func : callable, function to fit the data
    data_color : str, color for the data points
    fit_color : str, color for the fitted curve
    legend_loc : str, location for the legend
    to_print : bool, whether to print function parameters on the plot

    Returns:
    popt : array, optimal values for the parameters
    pcov : 2D array, the estimated covariance of popt
    goodness of fit: R^2 and chi^2_red
    '''
    # Plot the data with error bars
    if add_legend: label_data='Data'
    else: label_data=None
    ax.errorbar(x, y, xerr=abs(dx), yerr=abs(dy), fmt='+',
                color=data_color,label=label_data, alpha=data_alpha)
    # fit with errors via curve fit
    popt,pcov = scipy.optimize.curve_fit(fit_func, x, y,
            p0=None, sigma=np.sqrt(dx**2+dy**2), absolute_sigma=True)
    ax.plot(x,(fit_func(x,*popt)),color=fit_color,
            label='fit')#fit_func(x,*popt))

    # goodness of fit
    ss_total = np.sum((y - np.mean(y)) ** 2)
    line_of_best_fit = popt[0] * x + popt[1] # fit_func(x, *popt)
    residuals = y - line_of_best_fit
    ss_residual = np.sum(residuals**2)
    chi_squared_red = np.sum(residuals**2 / (dy**2+dx**2))/(len(x)-3)
    chi_sr_std = np.sqrt(2/(len(x)-3))
    r_squared = 1 - (ss_residual / ss_total)
    # need to fix printing of fitted values and goodness of fit for non-linear functions:
    # ax.plot(x,popt[0]*x+popt[1],'-',color=fit_color,
    #          label=rf'$\chi^{2}={chi_squared_red:.2f}, R^{2}: {r_squared:.2f}$,\
    #                  y = {popt[0]:.2f}x + {popt[1]:.2f} ')

    # upper =(popt[0]+np.diag(pcov)[0])*x+popt[1]+np.diag(pcov)[1]
    # lower = (popt[0]-np.diag(pcov)[0])*x+popt[1]-np.diag(pcov)[1]
    # ax.fill_between(np.arange(min(x),max(x),(max(x)-min(x))/len(x)),lower,upper)
    if add_legend: ax.legend(loc=legend_loc)
    print(rf'{popt=},{pcov=},{chi_squared_red=:.2e},{chi_sr_std=:.2e},{r_squared=:.2e}')
    return popt,pcov#,r_squared,chi_squared_red
#%% 44 get image resolution (can change other data type)

def get_image_resolution(image_path):
    """Get the resolution of an image."""
    with PIL.Image.open(image_path) as img:
        width, height = img.size
    return width, height

def extract_image_resolution(folder_path): # can change to get different data
    """Extract image data (resolution) from a folder."""
    data = []
    # Iterate over each file in the folder
    for file_name in os.listdir(folder_path):
        file_path = os.path.join(folder_path, file_name)
        # Check if the file is an image (you may need to adjust the list of supported extensions)
        if file_name.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
            # Get the resolution of the image
            resolution = get_image_resolution(file_path)
            # Extract the last part of the image file name (without extension) as the first column
            name = os.path.splitext(file_name)[0].split('_')[-1]
            # Append data to the list
            data.append({'Name': name, 'Width': resolution[0], 'Height': resolution[1]})
    # Create a DataFrame from the list of dictionaries
    df = pd.DataFrame(data)
    return df
#%% 45 write pandas dataframe to exp folder
def pdDF2xl(path,df):
    '''write pandas dataframe to exp folder'''

    # Check if the Excel file already exists
    try:
        # Load the existing Excel file
        book = load_workbook(path)

        # Create a writer object from the loaded Excel file
        writer = pd.ExcelWriter(path, engine='openpyxl')
        writer.book = book

        # Add the DataFrame to the existing Excel file (append to the first sheet)
        df.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)

        # Save the changes
        writer.save()
        writer.close()
    except FileNotFoundError:
        # If the file doesn't exist, create a new Excel file
        df.to_excel(path, index=False)

    print(f'Data written to {path}.')
#%% write function source code
# import inspect
# lines = inspect.getsource(foo)
# print(lines)
#%%